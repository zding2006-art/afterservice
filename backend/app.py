# -*- coding: utf-8 -*-
"""Flask API — 售后数据分析后台（V1.1 多语言 + 登录）"""

from flask import Flask, request, jsonify, send_file, send_from_directory, session, redirect, url_for
from flask_cors import CORS
from functools import wraps
import os
import sys
import json
import tempfile
import hashlib
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from analyzer import run_all_analysis
from db import init_db, save_monthly, get_all_months, get_month_detail, get_comparison_data, delete_month, _to_native
from insights import generate_insights


# ── PyInstaller 冻结环境路径解析 ──────────────────────────────
IS_FROZEN = getattr(sys, "frozen", False)

def _root_dir():
    """返回 exe / 源码所在目录（打包前后一致）"""
    if IS_FROZEN:
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


# 静态文件目录
_STATIC_DIR = os.path.join(_root_dir(), "frontend")
if not os.path.isdir(_STATIC_DIR):
    # 打包后 frontend 可能在 backend 旁边的源码目录
    alt = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "frontend")
    if os.path.isdir(alt):
        _STATIC_DIR = os.path.abspath(alt)

app = Flask(__name__, static_folder=_STATIC_DIR, static_url_path="")
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 0

# ── 登录配置 ──────────────────────────────────────────────
# 密码通过环境变量 SHOUHOU_PASSWORD 设置，默认值仅用于本地开发
_DEFAULT_PASSWORD = "waterx2026"
_ACCESS_PASSWORD = os.environ.get("SHOUHOU_PASSWORD", _DEFAULT_PASSWORD)
app.secret_key = os.environ.get("SHOUHOU_SECRET", "shouhou-analyzer-secret-2026")
_PORT = int(os.environ.get("PORT", 5859))

# ── CORS 配置 ──────────────────────────────────────────────
# CORS_ORIGINS 环境变量：逗号分隔的允许来源列表
# 设为 "*" 可允许所有来源（调试用）
_cors_origins_env = os.environ.get("CORS_ORIGINS", "")
if _cors_origins_env.strip() == "*":
    CORS(app, supports_credentials=False, origins="*")
elif _cors_origins_env.strip():
    CORS(app, supports_credentials=True, origins=[o.strip() for o in _cors_origins_env.split(",")])
else:
    _default_origins = [
        "http://localhost:5859", "http://127.0.0.1:5859",
        "http://localhost:80", "http://127.0.0.1:80",
        "http://localhost", "http://127.0.0.1",
    ]
    CORS(app, supports_credentials=True, origins=_default_origins)


def login_required(f):
    """登录验证装饰器：API 返回 401，页面重定向到登录页"""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("logged_in"):
            # 判断是 API 请求还是页面请求
            if request.path.startswith("/api/"):
                return jsonify({"error": "unauthorized", "redirect": "/login"}), 401
            return redirect("/login")
        return f(*args, **kwargs)
    return decorated


# ── 语言参数提取 ──────────────────────────────────────────
def _get_lang():
    """从请求中提取语言参数"""
    lang = request.args.get("lang") or request.headers.get("X-Lang") or "zh-CN"
    if lang not in ("zh-CN", "en", "ja"):
        lang = "zh-CN"
    return lang

UPLOAD_FOLDER = os.path.join(_root_dir(), "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# 数据库路径：打包后统一放在 exe 所在目录的 backend/ 下
import db
_BACKEND_DIR = os.path.join(_root_dir(), "backend")
_db_path = os.path.join(_BACKEND_DIR, "after_sales.db")
if IS_FROZEN:
    os.makedirs(_BACKEND_DIR, exist_ok=True)
    db.DB_PATH = _db_path

init_db()


# ── 登录页面 ──────────────────────────────────────────────
@app.route("/login", methods=["GET"])
def login_page():
    """登录页面 —— 与 index.html 是同一个文件，前端处理未登录状态"""
    return send_from_directory(app.static_folder, "index.html")


@app.route("/api/login", methods=["POST"])
def api_login():
    """登录 API"""
    data = request.get_json() or {}
    password = data.get("password", "")
    if password == _ACCESS_PASSWORD:
        session["logged_in"] = True
        session.permanent = True
        return jsonify({"success": True})
    return jsonify({"success": False, "error": "密码错误"}), 401


@app.route("/api/logout", methods=["POST"])
def api_logout():
    """退出登录"""
    session.clear()
    return jsonify({"success": True})


@app.route("/api/check-auth", methods=["GET"])
def check_auth():
    """检查登录状态"""
    return jsonify({"logged_in": bool(session.get("logged_in"))})


@app.route("/")
def index():
    return send_from_directory(app.static_folder, "index.html")


def _sheet_title_style(ws, title, row=1, merge_end="H"):
    """设置 sheet 标题样式"""
    ws.merge_cells(f"A{row}:{merge_end}{row}")
    cell = ws[f"A{row}"]
    cell.value = title
    cell.font = Font(name="Arial", size=14, bold=True, color="1F4E79")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30
    return row + 2


def _write_table(ws, start_row, data, headers, col_widths=None):
    """写入表格数据"""
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Arial", size=10)
    cell_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # 表头
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 数据行
    for ri, row_data in enumerate(data, start_row + 1):
        for ci, key in enumerate(headers, 1):
            val = row_data.get(key, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border

    # 列宽
    if col_widths:
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    return start_row + len(data) + 1


def _write_section_divider(ws, row, text):
    ws.merge_cells(f"A{row}:H{row}")
    cell = ws[f"A{row}"]
    cell.value = text
    cell.font = Font(name="Arial", size=11, bold=True, color="2E75B6")
    cell.fill = PatternFill("solid", fgColor="D6E4F0")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    return row + 1


def _auto_width(ws, start_row, end_row, max_col):
    for ci in range(1, max_col + 1):
        max_len = 0
        for ri in range(start_row, end_row + 1):
            val = ws.cell(row=ri, column=ci).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 6, 40)


@app.route("/api/upload", methods=["POST"])
@login_required
def upload_file():
    """上传 Excel 并执行全部分析。自动检测多月份数据并拆分保存。"""
    if "file" not in request.files:
        return jsonify({"error": "未上传文件"}), 400
    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "文件名为空"}), 400

    lang = _get_lang()
    filepath = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(filepath)

    try:
        df, results, summary, is_multi, month_details = run_all_analysis(filepath, lang=lang)

        # 提前获取已有月份列表，避免循环内重复查询
        existing_months = get_all_months()

        def _month_exists(y, m):
            return any(ex["year"] == y and ex["month"] == m for ex in existing_months)

        if is_multi and month_details:
            # 多月份：逐月检查并标记状态
            saved = []
            already_exists = []  # 已存在，返回给前端，由前端决定是否覆盖
            for md in month_details:
                if _month_exists(md["year"], md["month"]):
                    already_exists.append({
                        "year": md["year"],
                        "month": md["month"],
                        "month_label": md["month_label"],
                        "record_count": md["record_count"],
                        "summary": md["summary"],
                    })
                else:
                    raw_json = md.get("raw_data_json")
                    save_monthly(md["year"], md["month"], md["month_label"],
                                 md["summary"], md["results"], raw_json)
                    saved.append(md["month_label"])

            return jsonify({
                "success": True,
                "is_multi_month": True,
                "total_months_detected": len(month_details),
                "saved_months": saved,
                "already_exists_months": already_exists,   # 已存在的月份，前端可提示覆盖
                "summary": summary,
                "results": _to_native(results),
                "month_details": [{
                    "year": md["year"],
                    "month": md["month"],
                    "month_label": md["month_label"],
                    "record_count": md["record_count"],
                    "summary": md["summary"],
                } for md in month_details],
                # 传递各月完整结果供前端覆盖时使用（含 raw_data_json 供导出用）
                "_month_results_for_overwrite": {
                    f"{md['year']}-{md['month']}": {
                        "summary": md["summary"],
                        "results": _to_native(md["results"]),
                        "raw_data_json": md.get("raw_data_json"),
                    } for md in month_details if _month_exists(md["year"], md["month"])
                }
            })
        else:
            # 单月：提取不一致信息（analyzer 已附加在 summary 中）
            filename_mismatch = summary.pop("filename_mismatch", False)
            filename_year = summary.pop("filename_year", summary["year"])
            filename_month = summary.pop("filename_month", summary["month"])

            from analyzer import _df_to_json
            raw_json = _df_to_json(df)

            # 如果文件名与数据日期不一致，先返回给前端确认，暂不保存
            if filename_mismatch:
                return jsonify({
                    "success": True,
                    "is_multi_month": False,
                    "already_exists": False,
                    "filename_mismatch": True,
                    "filename_year": filename_year,
                    "filename_month": filename_month,
                    "data_year": summary["year"],
                    "data_month": summary["month"],
                    "summary": summary,
                    "results": _to_native(results),
                    "raw_data_json": raw_json,
                })

            # 无不一致：检查是否已存在
            already = _month_exists(summary["year"], summary["month"])
            if already:
                # 返回 already_exists=True，由前端询问是否覆盖
                return jsonify({
                    "success": True,
                    "is_multi_month": False,
                    "already_exists": True,
                    "summary": summary,
                    "results": _to_native(results),
                    "raw_data_json": raw_json,
                })
            else:
                save_monthly(summary["year"], summary["month"], summary["month_label"], summary, results, raw_json)
                return jsonify({
                    "success": True,
                    "is_multi_month": False,
                    "already_exists": False,
                    "summary": summary,
                    "results": _to_native(results)
                })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/overwrite", methods=["POST"])
@login_required
def overwrite_month():
    """覆盖保存指定月份（在前端确认后调用）"""
    payload = request.get_json()
    if not payload:
        return jsonify({"error": "无数据"}), 400
    year = payload.get("year")
    month = payload.get("month")
    month_label = payload.get("month_label")
    summary = payload.get("summary")
    results = payload.get("results")
    raw_data_json = payload.get("raw_data_json")
    if not all([year, month, month_label, summary, results]):
        return jsonify({"error": "参数不完整"}), 400
    try:
        # 先删除旧数据，再保存新数据
        existing = get_all_months()
        for m in existing:
            if m["year"] == year and m["month"] == month:
                delete_month(m["id"])
                break
        save_monthly(year, month, month_label, summary, results, raw_data_json)
        return jsonify({"success": True, "month_label": month_label})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/months", methods=["GET"])
@login_required
def list_months():
    months = get_all_months()
    return jsonify({"months": months})


@app.route("/api/months/<int:month_id>", methods=["GET"])
@login_required
def get_month(month_id):
    detail = get_month_detail(month_id)
    if not detail:
        return jsonify({"error": "未找到"}), 404
    return jsonify(detail)


@app.route("/api/months/<int:month_id>", methods=["DELETE"])
@login_required
def del_month(month_id):
    delete_month(month_id)
    return jsonify({"success": True})


@app.route("/api/comparison", methods=["GET"])
@login_required
def comparison():
    data = get_comparison_data()
    return jsonify(data)


@app.route("/api/export/<int:month_id>", methods=["GET"])
@login_required
def export_excel(month_id):
    """导出单月分析结果为 Excel"""
    detail = get_month_detail(month_id)
    if not detail:
        return jsonify({"error": "未找到"}), 404

    results = detail["analysis_data"]
    month_label = detail["month_label"]

    wb = Workbook()
    ws = wb.active
    ws.title = "分析汇总"

    row = _sheet_title_style(ws, f"{month_label} 售后数据分析报告", merge_end="G")

    # 汇总指标
    row = _write_section_divider(ws, row, "一、月度汇总指标")
    summary_data = [{
        "总记录数": detail["total_records"],
        "总费用(元)": detail["total_cost"],
        "保内件数": detail["in_warranty_count"],
        "保外件数": detail["out_warranty_count"],
        "盖机件数": detail["cover_count"],
        "整机件数": detail["whole_count"],
    }]
    row = _write_table(ws, row, summary_data,
                       ["总记录数", "总费用(元)", "保内件数", "保外件数", "盖机件数", "整机件数"],
                       [12, 14, 12, 12, 12, 12])

    # 各分析表单
    section_map = {
        "analysis_1": ("二、保修期内/外费用汇总", ["保内保外", "盖机整机", "件数", "部品费合计", "运费合计", "师傅费用合计", "总费用合计"]),
        "analysis_2": ("三、各产品使用月数与售后情况", ["型号", "盖机整机", "次数", "平均使用月数", "最小使用月数", "最大使用月数", "部品费合计", "总费用合计"]),
        "analysis_3": ("四、盖机保内 TOP10", ["排名", "型号", "发生次数", "部品费合计", "运费合计", "师傅费用合计", "总费用合计"]),
        "analysis_4": ("五、保内盖机 TOP10 各部品", None),  # 嵌套结构，特殊处理
        "analysis_5": ("六、保内盖机 TOP10 各部品排序", None),
        "analysis_6": ("七、整机保内 TOP10", ["排名", "型号", "发生次数", "部品费合计", "运费合计", "师傅费用合计", "总费用合计"]),
        "analysis_7": ("八、保内整机 TOP10 各部品", None),
        "analysis_8": ("九、保内整机 TOP10 各部品排序", None),
        "analysis_9": ("十、保内座盖投诉", ["发生次数", "部品费合计", "运费合计", "总费用合计"]),
        "analysis_10": ("十一、保内座盖软胶垫投诉", ["发生次数", "部品费合计", "运费合计", "总费用合计"]),
        "analysis_11": ("十二、柔光灯投诉", ["发生次数", "部品费合计", "运费合计", "总费用合计"]),
        "analysis_12": ("十三、座盖故障使用月数分布", ["使用月数区间", "发生次数"]),
        "analysis_13": ("十四、所有产品使用月数汇总", ["保内保外", "盖机整机", "型号", "件数", "平均使用月数", "最小使用月数", "最大使用月数"] + [f"{m}个月" for m in range(1, 25)] + ["3年(25~36月)", "4年(37~48月)", "大于4年", "月数合计"]),
    }

    for key, (section_title, headers) in section_map.items():
        result = results.get(key)
        if not result or not result.get("data"):
            continue
        row = _write_section_divider(ws, row, section_title)

        if headers:
            row = _write_table(ws, row, result["data"], headers)
        else:
            # 嵌套结构：多个型号各有多条部品记录
            for model_data in result["data"]:
                sub_title = f"型号: {model_data['型号']} (总次数: {model_data['总次数']})"
                row = _write_section_divider(ws, row, sub_title)
                parts = model_data.get("TOP10部品", [])
                parts_headers = ["排名", "配件", "发生次数", "部品费合计", "总费用合计"] if "排名" in (parts[0] if parts else {}) else ["配件", "发生次数", "部品费合计", "总费用合计"]
                row = _write_table(ws, row, parts, parts_headers)

    # 增加"处理后数据清单" sheet
    raw_json = detail.get("raw_data_json")
    if raw_json:
        try:
            raw_data = json.loads(raw_json)
        except (json.JSONDecodeError, TypeError):
            raw_data = None
        if raw_data:
            ws2 = wb.create_sheet("处理后数据清单")
            row2 = _sheet_title_style(ws2, f"{month_label} 处理后数据清单", merge_end="K")
            # 动态获取所有列名（含原始列+计算列）
            if raw_data:
                actual_headers = list(raw_data[0].keys())
                row2 = _write_table(ws2, row2, raw_data, actual_headers)

    wb.save(temp_path := os.path.join(tempfile.gettempdir(), f"after_sales_{month_id}.xlsx"))
    return send_file(temp_path, as_attachment=True, download_name=f"{month_label}售后分析.xlsx")


@app.route("/api/export/comparison-html", methods=["GET"])
@login_required
def export_comparison_html():
    """导出多月对比结果为 HTML 报告"""
    data = get_comparison_data()
    details = data["details"]
    if not details:
        return jsonify({"error": "暂无数据"}), 404

    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    # 构建表格行
    table_rows = ""
    for i, d in enumerate(details):
        ratio = data["in_warranty_ratio"][i] if i < len(data["in_warranty_ratio"]) else 0
        cover_pct = data["cover_ratio"][i] if i < len(data["cover_ratio"]) else 0
        table_rows += f"""
            <tr>
                <td>{d['month_label']}</td>
                <td>{d['total_records']}</td>
                <td>¥{d['total_cost']:,.0f}</td>
                <td>{d['in_warranty_count']}</td>
                <td>{d['out_warranty_count']}</td>
                <td>{d['cover_count']}</td>
                <td>{d['whole_count']}</td>
                <td>{ratio}%</td>
                <td>{cover_pct}%</td>
            </tr>"""

    labels_js = json.dumps(data["labels"], ensure_ascii=False)
    records_js = json.dumps(data["total_records"])
    cost_js = json.dumps(data["total_cost"])
    warranty_js = json.dumps(data["in_warranty_ratio"])
    cover_js = json.dumps(data["cover_ratio"])
    cover_counts = json.dumps([d["cover_count"] for d in details])
    whole_counts = json.dumps([d["whole_count"] for d in details])

    # --- 推移图表 JS + HTML（如果有 trends 数据）---
    trends_js = ""
    trends_html = ""
    if data.get("trends"):
        t = data["trends"]
        colors = ["'#2E75B6'","'#e53e3e'","'#38a169'","'#d69e2e'","'#805ad5'","'#dd6b20'","'#319795'","'#d53f8c'"]

        def _tbl_html(tbl):
            if not tbl: return ""
            rows = "".join(
                f"<tr><td style='font-weight:600;text-align:left;'>{r['label']}</td>" +
                "".join(f"<td>{v or 0}</td>" for v in r['data']) + "</tr>"
                for r in tbl.get("rows", [])
            )
            cols = "".join(f"<th>{c}</th>" for c in tbl.get("columns", []))
            return f"<table><thead><tr>{cols}</tr></thead><tbody>{rows}</tbody></table>"

        def _make_chart(chart_id, chart_type, labels, datasets, label_key, data_key):
            ds = json.dumps([{
                "label": ds[label_key], "data": ds[data_key],
                "borderColor": colors[i % 8] if chart_type == "line" else colors[i % 8],
                "backgroundColor": "transparent" if chart_type == "line" else colors[i % 8],
                "tension": 0.3, "pointRadius": 3
            } for i, ds in enumerate(datasets)])
            return f"""new Chart(document.getElementById('{chart_id}'), {{
    type: '{chart_type}',
    data: {{ labels: {labels_js}, datasets: {ds} }},
    options: {{ responsive: true, plugins: {{ legend: {{ position: 'bottom', labels: {{ usePointStyle: true, padding: 20, font: {{ size: 11 }} }} }} }} }}
}});"""

        # 盖机 TOP10
        tr = t["cover_top10"]
        trends_js += _make_chart("chartCoverTop10", "line", data["labels"], tr["datasets"], "model", "data") + "\n"
        trends_html += f'<div class="card"><div class="card-title">🔧 保内盖机 TOP10 型号推移</div><div style="height:320px;"><canvas id="chartCoverTop10"></canvas></div><div style="overflow-x:auto;margin-top:16px;">{_tbl_html(tr.get("table"))}</div></div>\n'

        # 整机 TOP10
        tr = t["whole_top10"]
        trends_js += _make_chart("chartWholeTop10", "line", data["labels"], tr["datasets"], "model", "data") + "\n"
        trends_html += f'<div class="card"><div class="card-title">📦 保内整机 TOP10 型号推移</div><div style="height:320px;"><canvas id="chartWholeTop10"></canvas></div><div style="overflow-x:auto;margin-top:16px;">{_tbl_html(tr.get("table"))}</div></div>\n'

        # 盖机部品汇总
        tr = t["cover_parts"]
        trends_js += _make_chart("chartCoverParts", "line", data["labels"], tr["datasets"], "part", "data") + "\n"
        trends_html += f'<div class="card"><div class="card-title">🔩 保内盖机 TOP10 部品推移（汇总）</div><div style="height:320px;"><canvas id="chartCoverParts"></canvas></div><div style="overflow-x:auto;margin-top:16px;">{_tbl_html(tr.get("table"))}</div>'

        # 盖机部品明细
        cd = t.get("cover_parts_detail")
        if cd and cd.get("models"):
            trends_html += '<div style="margin-top:20px;padding-top:16px;border-top:2px solid #e2e8f0;"><div style="font-weight:700;color:#1F4E79;font-size:15px;margin-bottom:12px;">📊 各型号 TOP10 部品明细</div>'
            for mi in cd["models"]:
                trends_html += f'<h4 style="color:#2E75B6;margin:12px 0 6px;">型号：{mi["model"]}</h4><table><thead><tr><th>部品</th>{"".join(f"<th>{l}</th>" for l in cd["labels"])}</tr></thead><tbody>'
                for p in mi["parts"]:
                    trends_html += f"<tr><td style='font-weight:500;'>{p['name']}</td>" + "".join(f"<td>{v or 0}</td>" for v in p["data"]) + "</tr>"
                trends_html += "</tbody></table>"
            trends_html += "</div>"
        trends_html += "</div>\n"

        # 整机部品汇总
        tr = t["whole_parts"]
        trends_js += _make_chart("chartWholeParts", "line", data["labels"], tr["datasets"], "part", "data") + "\n"
        trends_html += f'<div class="card"><div class="card-title">⚙️ 保内整机 TOP10 部品推移（汇总）</div><div style="height:320px;"><canvas id="chartWholeParts"></canvas></div><div style="overflow-x:auto;margin-top:16px;">{_tbl_html(tr.get("table"))}</div>'

        # 整机部品明细
        wd = t.get("whole_parts_detail")
        if wd and wd.get("models"):
            trends_html += '<div style="margin-top:20px;padding-top:16px;border-top:2px solid #e2e8f0;"><div style="font-weight:700;color:#1F4E79;font-size:15px;margin-bottom:12px;">📊 各型号 TOP10 部品明细</div>'
            for mi in wd["models"]:
                trends_html += f'<h4 style="color:#2E75B6;margin:12px 0 6px;">型号：{mi["model"]}</h4><table><thead><tr><th>部品</th>{"".join(f"<th>{l}</th>" for l in wd["labels"])}</tr></thead><tbody>'
                for p in mi["parts"]:
                    trends_html += f"<tr><td style='font-weight:500;'>{p['name']}</td>" + "".join(f"<td>{v or 0}</td>" for v in p["data"]) + "</tr>"
                trends_html += "</tbody></table>"
            trends_html += "</div>"
        trends_html += "</div>\n"

        # 专项推移
        tr = t["special"]
        special_ds = json.dumps([{"label": ds["label"], "data": ds["data"],
            "borderColor": ["'#e53e3e'","'#d69e2e'","'#805ad5'"][i], "backgroundColor": "transparent",
            "tension": 0.3, "pointRadius": 4, "borderWidth": 2} for i, ds in enumerate(tr["datasets"])])
        trends_js += f"""new Chart(document.getElementById('chartSpecial'), {{
    type: 'line',
    data: {{ labels: {labels_js}, datasets: {special_ds} }},
    options: {{ responsive: true, plugins: {{ legend: {{ position: 'bottom', labels: {{ usePointStyle: true, padding: 20, font: {{ size: 11 }} }} }} }} }}
}});
"""
        trends_html += f'<div class="card"><div class="card-title">🎯 专项投诉推移</div><div style="height:280px;"><canvas id="chartSpecial"></canvas></div><div style="overflow-x:auto;margin-top:16px;">{_tbl_html(tr.get("table"))}</div></div>\n'

        # 座盖故障使用月数
        tr = t["seat_usage"]
        seat_ds = json.dumps([{"label": ds["bucket"], "data": ds["data"], "backgroundColor": colors[i % 8]}
                               for i, ds in enumerate(tr["datasets"])])
        trends_js += f"""new Chart(document.getElementById('chartSeatUsage'), {{
    type: 'bar',
    data: {{ labels: {labels_js}, datasets: {seat_ds} }},
    options: {{ responsive: true, plugins: {{ legend: {{ position: 'bottom', labels: {{ usePointStyle: true, padding: 20, font: {{ size: 11 }} }} }} }} }}
}});
"""
        trends_html += f'<div class="card"><div class="card-title">🪑 座盖故障使用月数推移</div><div style="height:320px;"><canvas id="chartSeatUsage"></canvas></div><div style="overflow-x:auto;margin-top:16px;">{_tbl_html(tr.get("table"))}</div></div>\n'

    # --- AI 智能分析结论 ---
    insights_html = ""
    try:
        insights = generate_insights(data)
        insights["generated_at"] = now
        sei = {"high": "🔴", "medium": "🟡", "low": "🟢"}
        slevel = {"warning": "⚠️ 预警", "info": "ℹ️ 关注"}

        if insights.get("key_findings"):
            sc = {"high": "#e53e3e", "medium": "#d69e2e", "low": "#38a169"}
            sbg = {"high": "#fff5f5", "medium": "#fffff0", "low": "#f0fff4"}
            items = "".join(
                '<div class="insight-item" style="border-left:3px solid %s;background:%s;padding:12px 16px;margin-bottom:8px;border-radius:4px;">'
                '<div style="font-weight:700;font-size:14px;margin-bottom:4px;">%s %s</div>'
                '<div style="font-size:13px;color:#4a5568;white-space:pre-line;">%s</div></div>'
                % (sc.get(f.get("severity", ""), "#718096"),
                   sbg.get(f.get("severity", ""), "#f7fafc"),
                   sei.get(f.get("severity", ""), ""), f["title"], f["detail"])
                for f in insights["key_findings"]
            )
            insights_html += '<div class="card" id="ai-findings"><div class="card-title">🔍 关键发现 (%d 项)</div>%s</div>' % (len(insights["key_findings"]), items)

        if insights.get("risk_alerts"):
            albg = {"warning": "#fffaf0", "info": "#ebf8ff"}
            albd = {"warning": "#fbd38d", "info": "#bee3f8"}
            items = "".join(
                '<div style="padding:10px 14px;margin-bottom:6px;background:%s;border-radius:6px;border:1px solid %s;">'
                '<div style="font-weight:600;font-size:13px;">%s %s</div>'
                '<div style="font-size:12px;color:#718096;margin-top:4px;">%s</div></div>'
                % (albg.get(a.get("level", ""), "#f7fafc"),
                   albd.get(a.get("level", ""), "#e2e8f0"),
                   slevel.get(a.get("level", ""), ""), a["title"], a["detail"])
                for a in insights["risk_alerts"]
            )
            insights_html += '<div class="card" id="ai-alerts"><div class="card-title">🚨 风险预警 (%d 项)</div>%s</div>' % (len(insights["risk_alerts"]), items)

        if insights.get("recommendations"):
            items = "".join(
                f'<div style="padding:12px 16px;margin-bottom:8px;background:#f0fff4;border-radius:6px;border:1px solid #c6f6d5;">'
                f'<div style="font-weight:700;font-size:14px;color:#276749;">✅ {r["title"]}</div>'
                f'<div style="font-size:13px;color:#4a5568;margin-top:4px;white-space:pre-line;">{r["detail"]}</div>'
                f'<div style="font-size:11px;color:#a0aec0;margin-top:6px;">📂 {r.get("category","")}</div></div>'
                for r in insights["recommendations"]
            )
            insights_html += f'<div class="card" id="ai-recommendations"><div class="card-title">💡 改进建议 ({len(insights["recommendations"])} 项)</div>{items}</div>'

        if insights.get("analysis_angles"):
            items = "".join(
                f'<div style="padding:10px 14px;margin-bottom:6px;background:#faf5ff;border-radius:6px;border:1px solid #e9d8fd;">'
                f'<div style="font-weight:600;font-size:13px;color:#6b46c1;">📐 {a["title"]}</div>'
                f'<div style="font-size:12px;color:#4a5568;margin-top:4px;">{a["description"]}</div>'
                f'<div style="font-size:11px;color:#a0aec0;margin-top:4px;">🔧 方法：{a.get("method","")}</div></div>'
                for a in insights["analysis_angles"]
            )
            insights_html += f'<div class="card" id="ai-angles"><div class="card-title">🔬 建议分析角度 ({len(insights["analysis_angles"])} 项)</div>{items}</div>'

        if insights.get("summary"):
            insights_html = (
                f'<div class="card" id="ai-summary" style="background:linear-gradient(135deg,#1a365d,#2a4365);color:#fff;">'
                f'<div class="card-title" style="color:#bee3f8;border-left-color:#63b3ed;">🤖 AI 智能分析结论</div>'
                f'<div style="font-size:15px;line-height:1.8;opacity:0.95;">{insights["summary"]}</div>'
                f'<div style="font-size:11px;opacity:0.6;margin-top:12px;">生成时间：{now} · 基于 {len(details)} 个月数据</div></div>'
            ) + insights_html
    except Exception:
        insights_html = ""

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>售后数据分析 — 多月对比报告</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ font-family: -apple-system, "Microsoft YaHei", "PingFang SC", sans-serif; background: #f0f4f8; color: #2d3748; padding: 24px; }}
.header {{ text-align: center; padding: 32px; background: linear-gradient(135deg, #1F4E79, #2E75B6); color: #fff; border-radius: 12px; margin-bottom: 24px; }}
.header h1 {{ font-size: 24px; margin-bottom: 8px; }}
.header p {{ opacity: 0.85; font-size: 14px; }}
.card {{ background: #fff; border-radius: 12px; padding: 24px; margin-bottom: 24px; box-shadow: 0 1px 3px rgba(0,0,0,0.08); }}
.card-title {{ font-size: 18px; font-weight: 700; color: #1F4E79; margin-bottom: 16px; padding-bottom: 8px; border-bottom: 2px solid #e2e8f0; }}
.charts-row {{ display: grid; grid-template-columns: 1fr 1fr; gap: 24px; margin-bottom: 16px; }}
.charts-row canvas {{ max-height: 300px; }}
table {{ width: 100%; border-collapse: collapse; font-size: 14px; }}
th {{ background: #1F4E79; color: #fff; padding: 10px 12px; text-align: center; font-weight: 600; }}
td {{ padding: 10px 12px; text-align: center; border-bottom: 1px solid #e2e8f0; }}
tr:nth-child(even) td {{ background: #f7fafc; }}
tr:hover td {{ background: #ebf4ff; }}
.footer {{ text-align: center; color: #a0aec0; font-size: 12px; margin-top: 24px; }}
@media print {{ body {{ background: #fff; padding: 0; }} .card {{ box-shadow: none; break-inside: avoid; }} }}
</style>
</head>
<body>

<div class="header">
    <h1>📊 售后数据分析 — 多月对比报告</h1>
    <p>生成时间：{now} | 数据范围：{data["labels"][0]} ~ {data["labels"][-1]}（共 {len(details)} 个月）</p>
</div>

<div class="card">
    <div class="card-title">📈 趋势图表</div>
    <div class="charts-row">
        <div><h4 style="text-align:center;margin-bottom:8px;color:#4a5568;">总记录数趋势</h4><canvas id="chartRecords"></canvas></div>
        <div><h4 style="text-align:center;margin-bottom:8px;color:#4a5568;">总费用趋势 (¥)</h4><canvas id="chartCost"></canvas></div>
    </div>
    <div class="charts-row">
        <div><h4 style="text-align:center;margin-bottom:8px;color:#4a5568;">保内占比趋势 (%)</h4><canvas id="chartWarranty"></canvas></div>
        <div><h4 style="text-align:center;margin-bottom:8px;color:#4a5568;">盖机 / 整机对比</h4><canvas id="chartUnitType"></canvas></div>
    </div>
</div>

<div class="card">
    <div class="card-title">📋 历月数据汇总</div>
    <table>
        <thead><tr>
            <th>月份</th><th>总记录数</th><th>总费用</th><th>保内件数</th><th>保外件数</th>
            <th>盖机件数</th><th>整机件数</th><th>保内占比</th><th>盖机占比</th>
        </tr></thead>
        <tbody>{table_rows}</tbody>
    </table>
</div>

{insights_html}

{trends_html}

<div class="footer">
    <p>株式会社 Water X Technologies · 售后数据分析系统 · 自动生成</p>
</div>

<script>
const labels = {labels_js};
new Chart(document.getElementById('chartRecords'), {{
    type: 'bar',
    data: {{ labels, datasets: [{{ label: '总记录数', data: {records_js}, backgroundColor: '#2E75B6' }}] }},
    options: {{ responsive: true, plugins: {{ legend: {{ display: false }} }} }}
}});
new Chart(document.getElementById('chartCost'), {{
    type: 'line',
    data: {{ labels, datasets: [{{ label: '总费用(元)', data: {cost_js}, borderColor: '#e53e3e', backgroundColor: 'rgba(229,62,62,0.1)', tension: 0.3, fill: true }}] }},
    options: {{ responsive: true, plugins: {{ legend: {{ display: false }} }} }}
}});
new Chart(document.getElementById('chartWarranty'), {{
    type: 'line',
    data: {{ labels, datasets: [{{ label: '保内占比(%)', data: {warranty_js}, borderColor: '#38a169', backgroundColor: 'rgba(56,161,105,0.1)', tension: 0.3, fill: true }}] }},
    options: {{ responsive: true, plugins: {{ legend: {{ display: false }} }} }}
}});
new Chart(document.getElementById('chartUnitType'), {{
    type: 'bar',
    data: {{ labels, datasets: [
        {{ label: '盖机', data: {cover_counts}, backgroundColor: '#2E75B6' }},
        {{ label: '整机', data: {whole_counts}, backgroundColor: '#718096' }}
    ] }},
    options: {{ responsive: true }}
}});
{trends_js}
</script>
</body>
</html>"""

    temp_path = os.path.join(tempfile.gettempdir(), "after_sales_comparison.html")
    with open(temp_path, "w", encoding="utf-8") as f:
        f.write(html)
    return send_file(temp_path, as_attachment=True, download_name=f"售后多月对比报告_{datetime.now().strftime('%Y%m%d')}.html",
                     mimetype="text/html; charset=utf-8")


@app.route("/api/analysis/insights", methods=["GET"])
def ai_insights():
    """AI 智能分析 — 从多月对比数据中识别趋势、异常、生成结论与建议
    支持 lang 参数：zh / en / ja
    """
    lang = request.args.get("lang", "zh")
    if lang not in ("zh", "en", "ja"):
        lang = "zh"
    data = get_comparison_data()
    if not data.get("details") or len(data["details"]) < 2:
        # 按语言返回"数据不足"提示
        msgs = {
            "zh": "数据不足（需要至少 2 个月的数据），无法进行 AI 分析。",
            "en": "Insufficient data (at least 2 months required) for AI analysis.",
            "ja": "データが不足しています（最低2ヶ月のデータが必要です）。AI 分析はできません。",
        }
        return jsonify({
            "summary": msgs.get(lang, msgs["zh"]),
            "key_findings": [],
            "risk_alerts": [],
            "recommendations": [],
            "analysis_angles": [],
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        })
    insights = generate_insights(data, lang=lang)
    insights["generated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    return jsonify(insights)


@app.route("/api/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"})


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=_PORT, debug=not IS_FROZEN)
