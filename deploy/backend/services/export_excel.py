# -*- coding: utf-8 -*-
"""
services/export_excel.py — Excel 报告生成服务

职责：将单月分析结果（analysis_data + raw_data_json）渲染为 .xlsx 文件。
与 Flask、数据库无关，可单独测试。
"""

import json
import os
import tempfile

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── 样式常量 ──────────────────────────────────────────────

_HEADER_FONT      = Font(name="Arial", size=10, bold=True, color="FFFFFF")
_HEADER_FILL      = PatternFill("solid", fgColor="1F4E79")
_HEADER_ALIGN     = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_FONT        = Font(name="Arial", size=10)
_CELL_ALIGN       = Alignment(horizontal="center", vertical="center")
_THIN_BORDER      = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
_TITLE_FONT       = Font(name="Arial", size=14, bold=True, color="1F4E79")
_DIVIDER_FONT     = Font(name="Arial", size=11, bold=True, color="2E75B6")
_DIVIDER_FILL     = PatternFill("solid", fgColor="D6E4F0")


# ── 工具函数 ──────────────────────────────────────────────

def _sheet_title(ws, title: str, row: int = 1, merge_end: str = "H") -> int:
    """写入 sheet 标题，返回下一个可写行号。"""
    ws.merge_cells(f"A{row}:{merge_end}{row}")
    cell = ws[f"A{row}"]
    cell.value = title
    cell.font = _TITLE_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30
    return row + 2


def _section_divider(ws, row: int, text: str) -> int:
    """写入分区标题行，返回下一个可写行号。"""
    ws.merge_cells(f"A{row}:H{row}")
    cell = ws[f"A{row}"]
    cell.value = text
    cell.font = _DIVIDER_FONT
    cell.fill = _DIVIDER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    return row + 1


def _write_table(ws, start_row: int, data: list, headers: list,
                 col_widths: list = None) -> int:
    """将 data 写入表格，返回下一个可写行号。

    Args:
        ws:         目标 worksheet
        start_row:  表头起始行
        data:       行数据列表，每行是 dict
        headers:    列名列表（同时用作 dict key）
        col_widths: 可选列宽列表
    """
    # 表头
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=ci, value=h)
        cell.font   = _HEADER_FONT
        cell.fill   = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER

    # 数据行
    for ri, row_data in enumerate(data, start_row + 1):
        for ci, key in enumerate(headers, 1):
            cell = ws.cell(row=ri, column=ci, value=row_data.get(key, ""))
            cell.font      = _CELL_FONT
            cell.alignment = _CELL_ALIGN
            cell.border    = _THIN_BORDER

    # 列宽
    if col_widths:
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    return start_row + len(data) + 1


def _write_nested_parts(ws, row: int, model_rows: list, has_rank: bool = False) -> int:
    """写入嵌套的型号→部品表，返回下一个可写行号。"""
    parts_headers = (
        ["排名", "配件", "发生次数", "占比", "部品费合计", "总费用合计"]
        if has_rank
        else ["配件", "发生次数", "占比", "部品费合计", "总费用合计"]
    )
    for model_data in model_rows:
        sub_title = f"型号: {model_data['型号']} (总次数: {model_data['总次数']})"
        row = _section_divider(ws, row, sub_title)
        parts = model_data.get("TOP10部品", [])
        # 自动检测是否有排名列
        actual_headers = parts_headers if not parts else (
            ["排名", "配件", "发生次数", "占比", "部品费合计", "总费用合计"]
            if "排名" in parts[0] else
            ["配件", "发生次数", "占比", "部品费合计", "总费用合计"]
        )
        row = _write_table(ws, row, parts, actual_headers)
    return row


# ── 主导出函数 ──────────────────────────────────────────────

# 分析表段配置：key → (section_title, headers_or_None)
# None 表示嵌套结构，由 _write_nested_parts 处理
_SECTION_MAP = {
    "analysis_1":  ("二、保修期内/外费用汇总",
                    ["保内保外", "盖机整机", "件数", "部品费合计", "运费合计", "师傅费用合计", "总费用合计"]),
    "analysis_2":  ("三、各产品使用月数与售后情况",
                    ["型号", "盖机整机", "次数", "平均使用月数", "最小使用月数", "最大使用月数", "部品费合计", "总费用合计"]),
    "analysis_3":  ("四、盖机保内 TOP10",
                    ["排名", "型号", "发生次数", "部品费合计", "运费合计", "师傅费用合计", "总费用合计"]),
    "analysis_4":  ("五、保内盖机 TOP10 各部品",        None),
    "analysis_5":  ("六、保内盖机 TOP10 各部品排序",    None),
    "analysis_6":  ("七、整机保内 TOP10",
                    ["排名", "型号", "发生次数", "部品费合计", "运费合计", "师傅费用合计", "总费用合计"]),
    "analysis_7":  ("八、保内整机 TOP10 各部品",        None),
    "analysis_8":  ("九、保内整机 TOP10 各部品排序",    None),
    "analysis_9":  ("十、保内座盖投诉",
                    ["发生次数", "部品费合计", "运费合计", "总费用合计"]),
    "analysis_10": ("十一、保内座盖软胶垫投诉",
                    ["发生次数", "部品费合计", "运费合计", "总费用合计"]),
    "analysis_11": ("十二、柔光灯投诉",
                    ["发生次数", "部品费合计", "运费合计", "总费用合计"]),
    "analysis_12": ("十三、座盖故障使用月数分布",
                    ["使用月数区间", "发生次数"]),
    "analysis_13": ("十四、所有产品使用月数汇总",
                    ["保内保外", "盖机整机", "型号", "件数", "平均使用月数",
                     "最小使用月数", "最大使用月数"]
                    + [f"{m}个月" for m in range(1, 25)]
                    + ["3年(25~36月)", "4年(37~48月)", "大于4年", "月数合计"]),
}


def build_monthly_excel(month_label: str, detail: dict) -> str:
    """将单月分析结果生成 Excel 文件并返回临时文件路径。

    Args:
        month_label: 月份标签，如 "2025年9月"
        detail:      get_month_detail() 返回的完整记录
    Returns:
        临时文件的绝对路径
    """
    results = detail["analysis_data"]
    month_id = detail.get("id", "tmp")

    wb = Workbook()
    ws = wb.active
    ws.title = "分析汇总"

    # ── 主标题 ──
    row = _sheet_title(ws, f"{month_label} 售后数据分析报告", merge_end="G")

    # ── 汇总指标 ──
    row = _section_divider(ws, row, "一、月度汇总指标")
    summary_row = [{
        "总记录数":    detail["total_records"],
        "总费用(元)":  detail["total_cost"],
        "保内件数":    detail["in_warranty_count"],
        "保外件数":    detail["out_warranty_count"],
        "盖机件数":    detail["cover_count"],
        "整机件数":    detail["whole_count"],
    }]
    row = _write_table(ws, row, summary_row,
                       ["总记录数", "总费用(元)", "保内件数", "保外件数", "盖机件数", "整机件数"],
                       [12, 14, 12, 12, 12, 12])

    # ── 各分析表段 ──
    for key, (section_title, headers) in _SECTION_MAP.items():
        result = results.get(key)
        if not result or not result.get("data"):
            continue
        row = _section_divider(ws, row, section_title)
        if headers:
            row = _write_table(ws, row, result["data"], headers)
        else:
            row = _write_nested_parts(ws, row, result["data"])

    # ── 处理后数据清单（第二 sheet） ──
    raw_json = detail.get("raw_data_json")
    if raw_json:
        try:
            raw_data = json.loads(raw_json)
        except (json.JSONDecodeError, TypeError):
            raw_data = None
        if raw_data:
            ws2 = wb.create_sheet("处理后数据清单")
            row2 = _sheet_title(ws2, f"{month_label} 处理后数据清单", merge_end="K")
            actual_headers = list(raw_data[0].keys())
            _write_table(ws2, row2, raw_data, actual_headers)

    temp_path = os.path.join(tempfile.gettempdir(), f"after_sales_{month_id}.xlsx")
    wb.save(temp_path)
    return temp_path
